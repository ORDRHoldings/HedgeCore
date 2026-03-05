"""Bank Pack Excel generator using openpyxl."""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.schemas.results import CalculateResponse


def render_bank_pack_xlsx(result: CalculateResponse) -> bytes:
    wb = Workbook()

    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font_white = Font(bold=True, size=10, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    num_fmt_int = "#,##0"
    num_fmt_dec = "#,##0.00"
    num_fmt_rate = "#,##0.0000"

    # --- Sheet 1: Hedge Plan ---
    ws = wb.active
    ws.title = "Hedge Plan"

    headers = [
        "Bucket", "Confirmed MXN", "Forecast MXN", "Commercial Exp MXN",
        "Existing Hedges MXN", "Target Signed MXN", "Action MXN", "Direction",
        "Forward Rate", "Action USD", "Friction USD", "Suppressed",
        "Hedge Position MXN", "Residual MXN",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for r, b in enumerate(result.hedge_plan.buckets, 2):
        vals = [
            b.bucket, b.confirmed_flow_mxn, b.forecast_flow_mxn,
            b.commercial_exposure_mxn, b.existing_hedges_mxn,
            b.target_signed_mxn, b.action_mxn,
            b.action_direction or "-", b.forward_rate,
            b.action_usd, b.friction_usd,
            "Y" if b.suppressed else "N",
            b.hedge_position_mxn, b.residual_mxn,
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = thin_border
            if isinstance(v, float):
                if c == 9:  # forward rate
                    cell.number_format = num_fmt_rate
                elif c in (11,):  # friction
                    cell.number_format = num_fmt_dec
                else:
                    cell.number_format = num_fmt_int

    # Summary row
    sr = len(result.hedge_plan.buckets) + 2
    s = result.hedge_plan.summary
    ws.cell(row=sr, column=1, value="TOTAL").font = header_font
    summary_cols = {
        4: s.total_commercial_exposure_mxn,
        5: s.total_existing_hedges_mxn,
        7: s.total_action_mxn,
        10: s.total_action_usd,
        11: s.total_friction_usd,
        13: s.total_hedge_position_mxn,
        14: s.total_residual_mxn,
    }
    for c, v in summary_cols.items():
        cell = ws.cell(row=sr, column=c, value=v)
        cell.font = header_font
        cell.number_format = num_fmt_int if c != 11 else num_fmt_dec
        cell.border = thin_border

    # Auto-width
    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 18

    # --- Sheet 2: Scenarios ---
    ws2 = wb.create_sheet("Scenarios")
    scen_headers = ["Sigma", "Shocked Spot", "Total Unhedged USD", "Total Hedged USD", "Hedge Benefit USD"]
    for c, h in enumerate(scen_headers, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    for r, t in enumerate(result.scenario_results.totals, 2):
        vals = [t.sigma, t.shocked_spot, t.total_unhedged_usd, t.total_hedged_usd, t.total_hedge_benefit_usd]
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(row=r, column=c, value=v)
            cell.border = thin_border
            if c == 1:
                cell.number_format = "0%"
            elif c == 2:
                cell.number_format = num_fmt_rate
            else:
                cell.number_format = num_fmt_dec

    for c in range(1, len(scen_headers) + 1):
        ws2.column_dimensions[get_column_letter(c)].width = 22

    # --- Sheet 3: Audit ---
    ws3 = wb.create_sheet("Audit")
    audit_data = [
        ("Run ID", result.run_id),
        ("Timestamp", str(result.run_envelope.timestamp)),
        ("Engine Version", result.run_envelope.engine_version),
        ("Inputs Hash", result.run_envelope.inputs_hash),
        ("Outputs Hash", result.run_envelope.outputs_hash),
        ("Trades Hash", result.run_envelope.trades_hash),
        ("Hedges Hash", result.run_envelope.hedges_hash),
        ("Market Hash", result.run_envelope.market_hash),
        ("Policy Hash", result.run_envelope.policy_hash),
    ]
    for r, (label, value) in enumerate(audit_data, 1):
        ws3.cell(row=r, column=1, value=label).font = header_font
        ws3.cell(row=r, column=2, value=value)

    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["B"].width = 70

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
