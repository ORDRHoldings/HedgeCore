"""
tests/test_exports_units.py

Unit tests for app.exports_v1 — Excel, PDF, and ZIP builders.
Tests pure functions using mock CalculateResponse objects.
"""
from __future__ import annotations

import io
import json
import zipfile
from datetime import datetime, timezone
from unittest.mock import MagicMock

import pytest


# ---------------------------------------------------------------------------
# Build a mock CalculateResponse with realistic structure
# ---------------------------------------------------------------------------

def _mock_bucket(**overrides):
    b = MagicMock()
    b.bucket = overrides.get("bucket", "2026-01")
    b.confirmed_flow_mxn = overrides.get("confirmed_flow_mxn", -100_000.0)
    b.forecast_flow_mxn = overrides.get("forecast_flow_mxn", -50_000.0)
    b.commercial_exposure_mxn = overrides.get("commercial_exposure_mxn", -150_000.0)
    b.existing_hedges_mxn = overrides.get("existing_hedges_mxn", 50_000.0)
    b.target_signed_mxn = overrides.get("target_signed_mxn", 142_500.0)
    b.action_mxn = overrides.get("action_mxn", 92_500.0)
    b.action_direction = overrides.get("action_direction", "SELL_MXN_BUY_USD")
    b.forward_rate = overrides.get("forward_rate", 17.5000)
    b.action_usd = overrides.get("action_usd", 5_285.71)
    b.friction_usd = overrides.get("friction_usd", 8.57)
    b.suppressed = overrides.get("suppressed", False)
    b.hedge_position_mxn = overrides.get("hedge_position_mxn", 142_500.0)
    b.residual_mxn = overrides.get("residual_mxn", -7_500.0)
    return b


def _mock_summary():
    s = MagicMock()
    s.total_commercial_exposure_mxn = -300_000.0
    s.total_existing_hedges_mxn = 100_000.0
    s.total_action_mxn = 185_000.0
    s.total_action_usd = 10_571.42
    s.total_friction_usd = 17.14
    s.total_hedge_position_mxn = 285_000.0
    s.total_residual_mxn = -15_000.0
    return s


def _mock_scenario_total(**overrides):
    t = MagicMock()
    t.sigma = overrides.get("sigma", -0.10)
    t.shocked_spot = overrides.get("shocked_spot", 15.7500)
    t.total_unhedged_usd = overrides.get("total_unhedged_usd", -19_047.62)
    t.total_hedged_usd = overrides.get("total_hedged_usd", -16_285.71)
    t.total_hedge_benefit_usd = overrides.get("total_hedge_benefit_usd", 2_761.91)
    return t


def _mock_result():
    result = MagicMock()
    result.run_id = "test-run-001"

    # Run envelope
    result.run_envelope.timestamp = datetime(2026, 1, 15, 12, 0, 0, tzinfo=timezone.utc)
    result.run_envelope.engine_version = "1.0.0"
    result.run_envelope.inputs_hash = "a" * 64
    result.run_envelope.outputs_hash = "b" * 64
    result.run_envelope.trades_hash = "c" * 64
    result.run_envelope.hedges_hash = "d" * 64
    result.run_envelope.market_hash = "e" * 64
    result.run_envelope.policy_hash = "f" * 64

    # Buckets
    buckets = [_mock_bucket(bucket="2026-01"), _mock_bucket(bucket="2026-02")]
    result.hedge_plan.buckets = buckets
    result.hedge_plan.summary = _mock_summary()

    # Scenarios
    result.scenario_results.totals = [
        _mock_scenario_total(sigma=-0.10, shocked_spot=15.75),
        _mock_scenario_total(sigma=-0.05, shocked_spot=16.625),
        _mock_scenario_total(sigma=0.05, shocked_spot=18.375),
        _mock_scenario_total(sigma=0.10, shocked_spot=19.25),
    ]

    # Validation/trace for ZIP builder
    result.validation_report.model_dump.return_value = {"status": "PASS", "warnings": []}
    result.hedge_plan.model_dump.return_value = {"buckets": [], "summary": {}}
    result.scenario_results.model_dump.return_value = {"totals": []}
    result.run_envelope.model_dump.return_value = {"run_id": "test-run-001"}
    result.trace_lite.model_dump.return_value = {"steps": []}

    return result


# ---------------------------------------------------------------------------
# Excel Builder Tests
# ---------------------------------------------------------------------------

class TestExcelBuilder:
    def test_renders_bytes(self):
        from app.exports_v1.excel_builder import render_bank_pack_xlsx
        result = _mock_result()
        output = render_bank_pack_xlsx(result)
        assert isinstance(output, bytes)
        assert len(output) > 0

    def test_valid_xlsx_format(self):
        from app.exports_v1.excel_builder import render_bank_pack_xlsx
        from openpyxl import load_workbook
        result = _mock_result()
        output = render_bank_pack_xlsx(result)
        wb = load_workbook(io.BytesIO(output))
        assert len(wb.sheetnames) == 3
        assert "Hedge Plan" in wb.sheetnames
        assert "Scenarios" in wb.sheetnames
        assert "Audit" in wb.sheetnames

    def test_hedge_plan_sheet_has_headers(self):
        from app.exports_v1.excel_builder import render_bank_pack_xlsx
        from openpyxl import load_workbook
        result = _mock_result()
        output = render_bank_pack_xlsx(result)
        wb = load_workbook(io.BytesIO(output))
        ws = wb["Hedge Plan"]
        assert ws.cell(1, 1).value == "Bucket"
        assert ws.cell(1, 14).value == "Residual MXN"

    def test_hedge_plan_data_rows(self):
        from app.exports_v1.excel_builder import render_bank_pack_xlsx
        from openpyxl import load_workbook
        result = _mock_result()
        output = render_bank_pack_xlsx(result)
        wb = load_workbook(io.BytesIO(output))
        ws = wb["Hedge Plan"]
        assert ws.cell(2, 1).value == "2026-01"
        assert ws.cell(3, 1).value == "2026-02"

    def test_scenarios_sheet(self):
        from app.exports_v1.excel_builder import render_bank_pack_xlsx
        from openpyxl import load_workbook
        result = _mock_result()
        output = render_bank_pack_xlsx(result)
        wb = load_workbook(io.BytesIO(output))
        ws = wb["Scenarios"]
        assert ws.cell(1, 1).value == "Sigma"
        # 4 scenario rows
        assert ws.cell(2, 1).value == pytest.approx(-0.10)
        assert ws.cell(5, 1).value == pytest.approx(0.10)

    def test_audit_sheet_has_run_id(self):
        from app.exports_v1.excel_builder import render_bank_pack_xlsx
        from openpyxl import load_workbook
        result = _mock_result()
        output = render_bank_pack_xlsx(result)
        wb = load_workbook(io.BytesIO(output))
        ws = wb["Audit"]
        assert ws.cell(1, 1).value == "Run ID"
        assert ws.cell(1, 2).value == "test-run-001"

    def test_suppressed_bucket_shows_Y(self):
        from app.exports_v1.excel_builder import render_bank_pack_xlsx
        from openpyxl import load_workbook
        result = _mock_result()
        result.hedge_plan.buckets[0].suppressed = True
        output = render_bank_pack_xlsx(result)
        wb = load_workbook(io.BytesIO(output))
        ws = wb["Hedge Plan"]
        assert ws.cell(2, 12).value == "Y"

    def test_no_direction_shows_dash(self):
        from app.exports_v1.excel_builder import render_bank_pack_xlsx
        from openpyxl import load_workbook
        result = _mock_result()
        result.hedge_plan.buckets[0].action_direction = None
        output = render_bank_pack_xlsx(result)
        wb = load_workbook(io.BytesIO(output))
        ws = wb["Hedge Plan"]
        assert ws.cell(2, 8).value == "-"


# ---------------------------------------------------------------------------
# PDF Builder Tests
# ---------------------------------------------------------------------------

class TestPdfBuilder:
    def test_renders_bytes(self):
        from app.exports_v1.pdf_builder import render_bank_pack_pdf
        result = _mock_result()
        output = render_bank_pack_pdf(result)
        assert isinstance(output, bytes)
        assert len(output) > 100

    def test_starts_with_pdf_header(self):
        from app.exports_v1.pdf_builder import render_bank_pack_pdf
        result = _mock_result()
        output = render_bank_pack_pdf(result)
        assert output[:5] == b"%PDF-"

    def test_empty_buckets_still_works(self):
        from app.exports_v1.pdf_builder import render_bank_pack_pdf
        result = _mock_result()
        result.hedge_plan.buckets = []
        result.scenario_results.totals = []
        output = render_bank_pack_pdf(result)
        assert output[:5] == b"%PDF-"

    def test_null_direction_handled(self):
        from app.exports_v1.pdf_builder import render_bank_pack_pdf
        result = _mock_result()
        result.hedge_plan.buckets[0].action_direction = None
        output = render_bank_pack_pdf(result)
        assert isinstance(output, bytes)


# ---------------------------------------------------------------------------
# ZIP Builder Tests
# ---------------------------------------------------------------------------

class TestZipBuilder:
    def test_renders_valid_zip(self):
        from app.exports_v1.zip_builder import build_audit_zip
        result = _mock_result()
        output = build_audit_zip(result)
        assert isinstance(output, bytes)
        zf = zipfile.ZipFile(io.BytesIO(output))
        assert zf.testzip() is None  # no corrupt files

    def test_zip_contains_expected_files(self):
        from app.exports_v1.zip_builder import build_audit_zip
        result = _mock_result()
        output = build_audit_zip(result)
        zf = zipfile.ZipFile(io.BytesIO(output))
        names = set(zf.namelist())
        assert "ValidationReport.json" in names
        assert "HedgePlan.json" in names
        assert "ScenarioResults.json" in names
        assert "RunEnvelope.json" in names
        assert "TraceLite.json" in names
        assert "ExposureLedger.csv" in names
        assert "HedgeInstruction.csv" in names
        assert "BankPack.pdf" in names
        assert "BankPack.xlsx" in names
        assert "ReadMe.txt" in names

    def test_readme_contains_run_id(self):
        from app.exports_v1.zip_builder import build_audit_zip
        result = _mock_result()
        output = build_audit_zip(result)
        zf = zipfile.ZipFile(io.BytesIO(output))
        readme = zf.read("ReadMe.txt").decode("utf-8")
        assert "test-run-001" in readme
        assert "DETERMINISM STATEMENT" in readme

    def test_json_artifacts_are_valid(self):
        from app.exports_v1.zip_builder import build_audit_zip
        result = _mock_result()
        output = build_audit_zip(result)
        zf = zipfile.ZipFile(io.BytesIO(output))
        for name in ["ValidationReport.json", "HedgePlan.json", "ScenarioResults.json"]:
            data = json.loads(zf.read(name))
            assert isinstance(data, dict)


# ---------------------------------------------------------------------------
# CSV Generators
# ---------------------------------------------------------------------------

class TestCsvGenerators:
    def test_exposure_ledger_csv(self):
        from app.exports_v1.zip_builder import generate_exposure_ledger_csv
        buckets = [_mock_bucket(bucket="2026-01"), _mock_bucket(bucket="2026-02")]
        csv = generate_exposure_ledger_csv(buckets)
        lines = csv.split("\n")
        assert len(lines) == 3  # header + 2 data rows
        assert lines[0].startswith("Bucket,")
        assert "2026-01" in lines[1]

    def test_hedge_instruction_csv(self):
        from app.exports_v1.zip_builder import generate_hedge_instruction_csv
        buckets = [_mock_bucket(), _mock_bucket(action_direction=None, suppressed=True)]
        csv = generate_hedge_instruction_csv(buckets)
        lines = csv.split("\n")
        assert len(lines) == 3
        assert "SELL_MXN_BUY_USD" in lines[1]
        assert "NONE" in lines[2]
        assert "TRUE" in lines[2]

    def test_exposure_ledger_empty(self):
        from app.exports_v1.zip_builder import generate_exposure_ledger_csv
        csv = generate_exposure_ledger_csv([])
        lines = csv.split("\n")
        assert len(lines) == 1  # header only

    def test_hedge_instruction_empty(self):
        from app.exports_v1.zip_builder import generate_hedge_instruction_csv
        csv = generate_hedge_instruction_csv([])
        lines = csv.split("\n")
        assert len(lines) == 1


# ---------------------------------------------------------------------------
# Market Snapshot Service — pure hash functions
# ---------------------------------------------------------------------------

class TestMarketSnapshotHash:
    def test_canonical_payload_deterministic(self):
        from app.services.market_snapshot_service import build_canonical_payload
        a = build_canonical_payload({"z": 1, "a": 2})
        b = build_canonical_payload({"a": 2, "z": 1})
        assert a == b

    def test_canonical_payload_compact(self):
        from app.services.market_snapshot_service import build_canonical_payload
        result = build_canonical_payload({"key": "value"})
        assert " " not in result
        assert result == '{"key":"value"}'

    def test_snapshot_hash_hex_64(self):
        from app.services.market_snapshot_service import build_snapshot_hash
        h = build_snapshot_hash('{"test":1}')
        assert isinstance(h, str)
        assert len(h) == 64
        assert all(c in "0123456789abcdef" for c in h)

    def test_snapshot_hash_deterministic(self):
        from app.services.market_snapshot_service import build_snapshot_hash
        a = build_snapshot_hash('{"x":1}')
        b = build_snapshot_hash('{"x":1}')
        assert a == b

    def test_different_payload_different_hash(self):
        from app.services.market_snapshot_service import (
            build_canonical_payload,
            build_snapshot_hash,
        )
        h1 = build_snapshot_hash(build_canonical_payload({"rate": 17.5}))
        h2 = build_snapshot_hash(build_canonical_payload({"rate": 18.0}))
        assert h1 != h2
