"""
Tests for backend/app/services/audit_lab_parsers.py

Covers:
  - Shared helpers (normalize_headers, parse_date, parse_float, row_canonical,
    row_hash, row_confidence)
  - CSV parser (parse_csv)
  - XLSX parser (parse_xlsx)
  - PDF parser (parse_pdf)  — mocked via pdfplumber
  - SWIFT MT300/MT320 parser (parse_swift_mt)
  - File-type detection (detect_file_type)
  - Field confidence scoring
"""
from __future__ import annotations

import hashlib
import io
import json
from datetime import date
from typing import Any
from unittest.mock import MagicMock, patch

import pytest

from app.services.audit_lab_parsers import (
    FIELD_ALIASES,
    _build_row,
    _parse_swift_amount,
    _split_swift_messages,
    _extract_swift_tags,
    detect_file_type,
    normalize_headers,
    parse_csv,
    parse_date,
    parse_float,
    parse_swift_mt,
    parse_xlsx,
    parse_pdf,
    row_canonical,
    row_confidence,
    row_hash,
)


# ═══════════════════════════════════════════════════════════════════════════════
# Shared helpers
# ═══════════════════════════════════════════════════════════════════════════════


class TestFieldAliases:
    def test_aliases_dict_is_not_empty(self):
        assert len(FIELD_ALIASES) >= 10

    def test_every_canonical_has_self_alias(self):
        """Each canonical name should be its own first alias."""
        for canonical, aliases in FIELD_ALIASES.items():
            assert canonical in aliases, f"{canonical} missing self-alias"


class TestNormalizeHeaders:
    def test_exact_match(self):
        headers = ["trade_date", "currency_sold", "amount_sold"]
        mapping = normalize_headers(headers)
        assert mapping["trade_date"] == "trade_date"
        assert mapping["currency_sold"] == "currency_sold"
        assert mapping["amount_sold"] == "amount_sold"

    def test_alias_match(self):
        headers = ["TradeDate", "sell_ccy", "buy_ccy", "from_amount", "to_amount"]
        mapping = normalize_headers(headers)
        assert mapping.get("trade_date") == "TradeDate"
        assert mapping.get("currency_sold") == "sell_ccy"
        assert mapping.get("currency_bought") == "buy_ccy"
        assert mapping.get("amount_sold") == "from_amount"
        assert mapping.get("amount_bought") == "to_amount"

    def test_unmapped_header_ignored(self):
        headers = ["trade_date", "random_column", "amount_sold"]
        mapping = normalize_headers(headers)
        assert "random_column" not in mapping
        assert "trade_date" in mapping

    def test_case_insensitive(self):
        headers = ["Trade Date", "SELL_CCY"]
        mapping = normalize_headers(headers)
        assert mapping.get("trade_date") == "Trade Date"
        assert mapping.get("currency_sold") == "SELL_CCY"

    def test_empty_headers(self):
        mapping = normalize_headers([])
        assert mapping == {}

    def test_whitespace_stripped(self):
        headers = ["  trade_date  ", " currency_sold "]
        mapping = normalize_headers(headers)
        assert "trade_date" in mapping
        assert "currency_sold" in mapping


class TestParseDate:
    @pytest.mark.parametrize("input_str,expected", [
        ("2025-06-15", date(2025, 6, 15)),
        ("15/06/2025", date(2025, 6, 15)),
        ("06/15/2025", date(2025, 6, 15)),
        ("15-06-2025", date(2025, 6, 15)),
    ])
    def test_valid_formats(self, input_str: str, expected: date):
        assert parse_date(input_str) == expected

    def test_none_input(self):
        assert parse_date(None) is None

    def test_empty_string(self):
        assert parse_date("") is None

    def test_whitespace_only(self):
        assert parse_date("   ") is None

    def test_invalid_format(self):
        assert parse_date("not-a-date") is None

    def test_strips_whitespace(self):
        assert parse_date("  2025-06-15  ") == date(2025, 6, 15)


class TestParseFloat:
    def test_simple_number(self):
        assert parse_float("1234.56") == 1234.56

    def test_comma_separated(self):
        assert parse_float("1,234,567.89") == 1234567.89

    def test_none_input(self):
        assert parse_float(None) is None

    def test_empty_string(self):
        assert parse_float("") is None

    def test_whitespace_only(self):
        assert parse_float("   ") is None

    def test_invalid_input(self):
        assert parse_float("abc") is None

    def test_strips_whitespace(self):
        assert parse_float("  42.0  ") == 42.0

    def test_negative(self):
        assert parse_float("-100.50") == -100.50


class TestRowCanonicalAndHash:
    def test_canonical_is_sorted_json(self):
        data = {"b": 2, "a": 1}
        result = row_canonical(data)
        assert result == '{"a": 1, "b": 2}'

    def test_canonical_handles_date(self):
        data = {"dt": date(2025, 1, 1)}
        result = row_canonical(data)
        assert "2025-01-01" in result

    def test_hash_is_sha256(self):
        data = {"x": 1}
        expected = hashlib.sha256(
            json.dumps(data, sort_keys=True, default=str).encode()
        ).hexdigest()
        assert row_hash(data) == expected
        assert len(row_hash(data)) == 64

    def test_hash_deterministic(self):
        data = {"z": 3, "a": 1}
        assert row_hash(data) == row_hash(data)

    def test_different_data_different_hash(self):
        assert row_hash({"a": 1}) != row_hash({"a": 2})


class TestRowConfidence:
    def test_no_warnings_returns_1(self):
        row = {"parse_warnings": []}
        assert row_confidence(row) == 1.0

    def test_string_warnings_ignored(self):
        row = {"parse_warnings": ["Row 0: missing trade_date"]}
        assert row_confidence(row) == 1.0

    def test_dict_warning_with_confidence(self):
        row = {"parse_warnings": [{"source": "pdf", "confidence": 0.7}]}
        assert row_confidence(row) == 0.7

    def test_multiple_confidences_returns_min(self):
        row = {
            "parse_warnings": [
                {"source": "pdf", "confidence": 0.9},
                {"source": "pdf", "confidence": 0.5},
            ]
        }
        assert row_confidence(row) == 0.5

    def test_missing_warnings_key(self):
        assert row_confidence({}) == 1.0


class TestBuildRow:
    def test_basic_row(self):
        row = _build_row(
            row_index=0,
            trade_date="2025-01-01",
            value_date="2025-01-03",
            currency_sold="USD",
            currency_bought="EUR",
            amount_sold=1000000.0,
            amount_bought=920000.0,
            counterparty="Citi",
            fee_amount=50.0,
            fee_currency="USD",
            reference="REF001",
            parse_warnings=[],
        )
        assert row["row_index"] == 0
        assert row["trade_date"] == "2025-01-01"
        assert row["currency_sold"] == "USD"
        assert row["effective_rate"] == pytest.approx(0.92)
        assert len(row["parse_warnings"]) == 0

    def test_missing_trade_date_warning(self):
        row = _build_row(
            row_index=5,
            trade_date=None,
            value_date=None,
            currency_sold="USD",
            currency_bought="EUR",
            amount_sold=100.0,
            amount_bought=90.0,
            counterparty=None,
            fee_amount=None,
            fee_currency=None,
            reference=None,
            parse_warnings=[],
        )
        str_warnings = [w for w in row["parse_warnings"] if isinstance(w, str)]
        assert any("missing trade_date" in w for w in str_warnings)

    def test_missing_currency_warning(self):
        row = _build_row(
            row_index=0,
            trade_date="2025-01-01",
            value_date=None,
            currency_sold=None,
            currency_bought=None,
            amount_sold=None,
            amount_bought=None,
            counterparty=None,
            fee_amount=None,
            fee_currency=None,
            reference=None,
            parse_warnings=[],
        )
        str_warnings = [w for w in row["parse_warnings"] if isinstance(w, str)]
        assert any("missing currency" in w for w in str_warnings)

    def test_zero_amount_sold_no_rate(self):
        row = _build_row(
            row_index=0,
            trade_date="2025-01-01",
            value_date=None,
            currency_sold="USD",
            currency_bought="EUR",
            amount_sold=0.0,
            amount_bought=100.0,
            counterparty=None,
            fee_amount=None,
            fee_currency=None,
            reference=None,
            parse_warnings=[],
        )
        assert row["effective_rate"] is None

    def test_preserves_incoming_warnings(self):
        row = _build_row(
            row_index=0,
            trade_date="2025-01-01",
            value_date=None,
            currency_sold="USD",
            currency_bought="EUR",
            amount_sold=100.0,
            amount_bought=90.0,
            counterparty=None,
            fee_amount=None,
            fee_currency=None,
            reference=None,
            parse_warnings=[{"source": "pdf", "confidence": 0.7}],
        )
        dict_warnings = [w for w in row["parse_warnings"] if isinstance(w, dict)]
        assert len(dict_warnings) >= 1
        assert dict_warnings[0]["confidence"] == 0.7


# ═══════════════════════════════════════════════════════════════════════════════
# File-type detection
# ═══════════════════════════════════════════════════════════════════════════════


class TestDetectFileType:
    def test_csv_by_extension(self):
        assert detect_file_type("trades.csv", b"a,b,c") == "csv"

    def test_tsv_by_extension(self):
        assert detect_file_type("data.tsv", b"a\tb\tc") == "csv"

    def test_xlsx_by_extension(self):
        assert detect_file_type("report.xlsx", b"PK\x03\x04") == "xlsx"

    def test_xls_by_extension(self):
        assert detect_file_type("report.xls", b"\xd0\xcf") == "xlsx"

    def test_pdf_by_extension(self):
        assert detect_file_type("statement.pdf", b"%PDF-1.4") == "pdf"

    def test_swift_by_extension(self):
        assert detect_file_type("msg.mt300", b":20:REF") == "swift"
        assert detect_file_type("msg.mt320", b":20:REF") == "swift"
        assert detect_file_type("msg.fin", b":20:REF") == "swift"
        assert detect_file_type("msg.swift", b":20:REF") == "swift"

    def test_pdf_by_magic(self):
        assert detect_file_type("unknown", b"%PDF-1.7 rest") == "pdf"

    def test_xlsx_by_magic(self):
        assert detect_file_type("unknown", b"PK\x03\x04rest") == "xlsx"

    def test_swift_by_content(self):
        content = b"{1:F01BANKBEBBAXXX0000000000}\n:20:REF123\n:32B:USD1000000,"
        assert detect_file_type("unknown", content) == "swift"

    def test_swift_by_tag_line(self):
        content = b":20:REF123\n:30T:20250615\n:32B:USD100000,\n"
        assert detect_file_type("unknown", content) == "swift"

    def test_csv_by_content(self):
        content = b"trade_date,currency_sold,amount\n2025-01-01,USD,100"
        assert detect_file_type("unknown_file", content) == "csv"

    def test_unknown_binary(self):
        assert detect_file_type("blob", bytes(range(256))) == "unknown"


# ═══════════════════════════════════════════════════════════════════════════════
# CSV Parser
# ═══════════════════════════════════════════════════════════════════════════════


class TestParseCSV:
    def _make_csv(self, headers: list[str], data_rows: list[list[str]]) -> bytes:
        buf = io.StringIO()
        buf.write(",".join(headers) + "\n")
        for row in data_rows:
            buf.write(",".join(row) + "\n")
        return buf.getvalue().encode("utf-8")

    def test_basic_parse(self):
        raw = self._make_csv(
            ["trade_date", "currency_sold", "currency_bought", "amount_sold", "amount_bought"],
            [["2025-06-15", "USD", "EUR", "1000000", "920000"]],
        )
        rows, warnings, pairs = parse_csv(raw)
        assert len(rows) == 1
        assert rows[0]["currency_sold"] == "USD"
        assert rows[0]["currency_bought"] == "EUR"
        assert rows[0]["amount_sold"] == 1000000.0
        assert rows[0]["effective_rate"] == pytest.approx(0.92)
        assert "USDEUR" in pairs

    def test_alias_headers(self):
        raw = self._make_csv(
            ["date", "sell_ccy", "buy_ccy", "from_amount", "to_amount"],
            [["2025-01-01", "GBP", "USD", "500000", "630000"]],
        )
        rows, warnings, pairs = parse_csv(raw)
        assert len(rows) == 1
        assert rows[0]["currency_sold"] == "GBP"
        assert "GBPUSD" in pairs

    def test_multiple_rows(self):
        raw = self._make_csv(
            ["trade_date", "currency_sold", "currency_bought", "amount_sold", "amount_bought"],
            [
                ["2025-01-01", "USD", "EUR", "100", "92"],
                ["2025-01-02", "USD", "GBP", "200", "156"],
                ["2025-01-03", "EUR", "JPY", "300", "48000"],
            ],
        )
        rows, warnings, pairs = parse_csv(raw)
        assert len(rows) == 3
        assert "USDEUR" in pairs
        assert "USDGBP" in pairs
        assert "EURJPY" in pairs

    def test_missing_fields_generate_warnings(self):
        raw = self._make_csv(
            ["currency_sold", "amount_sold"],
            [["USD", "1000"]],
        )
        rows, warnings, pairs = parse_csv(raw)
        assert len(rows) == 1
        str_warnings = [w for w in rows[0]["parse_warnings"] if isinstance(w, str)]
        assert any("missing trade_date" in w for w in str_warnings)
        assert any("missing currency" in w for w in str_warnings)

    def test_empty_csv_raises(self):
        with pytest.raises(ValueError, match="no headers"):
            parse_csv(b"")

    def test_bom_handled(self):
        raw = b"\xef\xbb\xbftrade_date,currency_sold,currency_bought,amount_sold,amount_bought\n2025-01-01,USD,EUR,100,92\n"
        rows, warnings, pairs = parse_csv(raw)
        assert len(rows) == 1
        assert rows[0]["trade_date"] == "2025-01-01"

    def test_comma_in_amounts(self):
        # Commas inside amounts require proper CSV quoting
        raw = (
            b'trade_date,currency_sold,currency_bought,amount_sold,amount_bought\n'
            b'2025-01-01,USD,EUR,"1,000,000","920,000"\n'
        )
        rows, _, _ = parse_csv(raw)
        assert rows[0]["amount_sold"] == 1000000.0
        assert rows[0]["amount_bought"] == 920000.0

    def test_optional_fields(self):
        raw = self._make_csv(
            ["trade_date", "currency_sold", "currency_bought", "amount_sold", "amount_bought",
             "counterparty", "fee_amount", "fee_currency", "reference"],
            [["2025-01-01", "USD", "EUR", "100", "92", "Citi", "50", "USD", "REF001"]],
        )
        rows, _, _ = parse_csv(raw)
        assert rows[0]["counterparty"] == "Citi"
        assert rows[0]["fee_amount"] == 50.0
        assert rows[0]["fee_currency"] == "USD"
        assert rows[0]["reference"] == "REF001"


# ═══════════════════════════════════════════════════════════════════════════════
# XLSX Parser
# ═══════════════════════════════════════════════════════════════════════════════


class TestParseXLSX:
    """Tests for the XLSX parser using real openpyxl workbooks."""

    @staticmethod
    def _make_xlsx(headers: list[str], data_rows: list[list[Any]]) -> bytes:
        """Create a minimal XLSX file in memory."""
        import openpyxl as xl

        wb = xl.Workbook()
        ws = wb.active
        ws.append(headers)
        for row in data_rows:
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_basic_parse(self):
        raw = self._make_xlsx(
            ["trade_date", "currency_sold", "currency_bought", "amount_sold", "amount_bought"],
            [["2025-06-15", "USD", "EUR", 1000000, 920000]],
        )
        rows, warnings, pairs = parse_xlsx(raw)
        assert len(rows) == 1
        assert rows[0]["currency_sold"] == "USD"
        assert rows[0]["amount_sold"] == 1000000.0
        assert "USDEUR" in pairs

    def test_alias_headers(self):
        raw = self._make_xlsx(
            ["date", "sell_ccy", "buy_ccy", "from_amount", "to_amount"],
            [["2025-01-01", "GBP", "USD", 500000, 630000]],
        )
        rows, _, pairs = parse_xlsx(raw)
        assert rows[0]["currency_sold"] == "GBP"
        assert "GBPUSD" in pairs

    def test_auto_detect_header_row_with_blank_prefix(self):
        """Header detection skips rows with fewer than 3 non-empty cells."""
        raw = self._make_xlsx(
            ["", ""],  # row 0 — only 2 cells, both empty
            [],
        )
        # Build a workbook with a real header on row 2
        import openpyxl as xl

        wb = xl.Workbook()
        ws = wb.active
        ws.append(["Report Title", None])                # row 1 (< 3 cells)
        ws.append([None, None])                           # row 2 (blank)
        ws.append(["trade_date", "currency_sold", "currency_bought", "amount_sold", "amount_bought"])  # row 3
        ws.append(["2025-01-01", "USD", "EUR", "100", "92"])
        buf = io.BytesIO()
        wb.save(buf)
        raw = buf.getvalue()

        rows, _, pairs = parse_xlsx(raw)
        assert len(rows) == 1
        assert rows[0]["currency_sold"] == "USD"

    def test_blank_rows_skipped(self):
        import openpyxl as xl

        wb = xl.Workbook()
        ws = wb.active
        ws.append(["trade_date", "currency_sold", "currency_bought", "amount_sold", "amount_bought"])
        ws.append(["2025-01-01", "USD", "EUR", "100", "92"])
        ws.append([None, None, None, None, None])  # blank row
        ws.append(["2025-01-02", "GBP", "JPY", "200", "30000"])
        buf = io.BytesIO()
        wb.save(buf)

        rows, _, pairs = parse_xlsx(buf.getvalue())
        assert len(rows) == 2

    def test_empty_sheet_raises(self):
        import openpyxl as xl

        wb = xl.Workbook()
        buf = io.BytesIO()
        wb.save(buf)

        with pytest.raises(ValueError, match="(empty|no.*header)"):
            parse_xlsx(buf.getvalue())

    def test_multiple_rows(self):
        raw = self._make_xlsx(
            ["trade_date", "currency_sold", "currency_bought", "amount_sold", "amount_bought"],
            [
                ["2025-01-01", "USD", "EUR", 100, 92],
                ["2025-01-02", "USD", "GBP", 200, 156],
            ],
        )
        rows, _, pairs = parse_xlsx(raw)
        assert len(rows) == 2
        assert "USDEUR" in pairs
        assert "USDGBP" in pairs

    def test_confidence_in_warnings(self):
        """When a mapped header exists but cell is empty, confidence < 1.0."""
        raw = self._make_xlsx(
            ["trade_date", "currency_sold", "currency_bought", "amount_sold", "amount_bought"],
            [["2025-01-01", "USD", "EUR", None, "92"]],
        )
        rows, _, _ = parse_xlsx(raw)
        # amount_sold is None -> mapped header present but empty -> confidence 0.8
        dict_warnings = [
            w for w in rows[0]["parse_warnings"]
            if isinstance(w, dict) and "confidence" in w
        ]
        assert len(dict_warnings) >= 1
        assert dict_warnings[0]["confidence"] <= 1.0

    def test_openpyxl_not_installed(self):
        with patch("app.services.audit_lab_parsers._HAS_OPENPYXL", False):
            with pytest.raises(ImportError, match="openpyxl"):
                parse_xlsx(b"fake")


# ═══════════════════════════════════════════════════════════════════════════════
# PDF Parser
# ═══════════════════════════════════════════════════════════════════════════════


class TestParsePDF:
    """Tests for the PDF parser using mocked pdfplumber."""

    def _mock_pdf(self, tables: list[list[list[str | None]]]) -> MagicMock:
        """Create a mock pdfplumber PDF object."""
        mock_pdf = MagicMock()
        pages = []
        for table in tables:
            page = MagicMock()
            page.extract_tables.return_value = [table]
            pages.append(page)
        mock_pdf.pages = pages
        mock_pdf.close = MagicMock()
        return mock_pdf

    @patch("app.services.audit_lab_parsers._HAS_PDFPLUMBER", True)
    @patch("app.services.audit_lab_parsers.pdfplumber")
    def test_basic_pdf_parse(self, mock_pdfplumber: MagicMock):
        table = [
            ["trade_date", "currency_sold", "currency_bought", "amount_sold", "amount_bought"],
            ["2025-06-15", "USD", "EUR", "1000000", "920000"],
        ]
        mock_pdfplumber.open.return_value = self._mock_pdf([table])

        rows, warnings, pairs = parse_pdf(b"fake-pdf-bytes")
        assert len(rows) == 1
        assert rows[0]["currency_sold"] == "USD"
        assert rows[0]["amount_sold"] == 1000000.0
        assert "USDEUR" in pairs

    @patch("app.services.audit_lab_parsers._HAS_PDFPLUMBER", True)
    @patch("app.services.audit_lab_parsers.pdfplumber")
    def test_pdf_confidence_in_warnings(self, mock_pdfplumber: MagicMock):
        table = [
            ["trade_date", "currency_sold", "currency_bought", "amount_sold", "amount_bought"],
            ["2025-06-15", "USD", "EUR", "1000000", "920000"],
        ]
        mock_pdfplumber.open.return_value = self._mock_pdf([table])

        rows, _, _ = parse_pdf(b"fake")
        dict_warnings = [
            w for w in rows[0]["parse_warnings"]
            if isinstance(w, dict) and "confidence" in w
        ]
        assert len(dict_warnings) >= 1
        conf = dict_warnings[0]["confidence"]
        assert 0.5 <= conf <= 0.9

    @patch("app.services.audit_lab_parsers._HAS_PDFPLUMBER", True)
    @patch("app.services.audit_lab_parsers.pdfplumber")
    def test_pdf_low_confidence_on_bad_amount(self, mock_pdfplumber: MagicMock):
        table = [
            ["trade_date", "currency_sold", "currency_bought", "amount_sold", "amount_bought"],
            ["2025-06-15", "USD", "EUR", "1OO0OOO", "920000"],  # OCR artifact in amount
        ]
        mock_pdfplumber.open.return_value = self._mock_pdf([table])

        rows, _, _ = parse_pdf(b"fake")
        conf = row_confidence(rows[0])
        assert conf <= 0.6  # degraded due to OCR artifact

    @patch("app.services.audit_lab_parsers._HAS_PDFPLUMBER", True)
    @patch("app.services.audit_lab_parsers.pdfplumber")
    def test_pdf_multi_page(self, mock_pdfplumber: MagicMock):
        table1 = [
            ["trade_date", "currency_sold", "currency_bought", "amount_sold", "amount_bought"],
            ["2025-01-01", "USD", "EUR", "100", "92"],
        ]
        table2 = [
            ["trade_date", "currency_sold", "currency_bought", "amount_sold", "amount_bought"],
            ["2025-01-02", "GBP", "JPY", "200", "30000"],
        ]
        mock_pdfplumber.open.return_value = self._mock_pdf([table1, table2])

        rows, _, pairs = parse_pdf(b"fake")
        assert len(rows) == 2
        assert "USDEUR" in pairs
        assert "GBPJPY" in pairs

    @patch("app.services.audit_lab_parsers._HAS_PDFPLUMBER", True)
    @patch("app.services.audit_lab_parsers.pdfplumber")
    def test_pdf_no_tables_raises(self, mock_pdfplumber: MagicMock):
        mock_pdf = MagicMock()
        page = MagicMock()
        page.extract_tables.return_value = []
        mock_pdf.pages = [page]
        mock_pdf.close = MagicMock()
        mock_pdfplumber.open.return_value = mock_pdf

        with pytest.raises(ValueError, match="No tables"):
            parse_pdf(b"fake")

    @patch("app.services.audit_lab_parsers._HAS_PDFPLUMBER", True)
    @patch("app.services.audit_lab_parsers.pdfplumber")
    def test_pdf_blank_rows_skipped(self, mock_pdfplumber: MagicMock):
        table = [
            ["trade_date", "currency_sold", "currency_bought", "amount_sold", "amount_bought"],
            ["2025-01-01", "USD", "EUR", "100", "92"],
            [None, None, None, None, None],
            ["2025-01-02", "GBP", "USD", "200", "252"],
        ]
        mock_pdfplumber.open.return_value = self._mock_pdf([table])

        rows, _, _ = parse_pdf(b"fake")
        assert len(rows) == 2

    def test_pdfplumber_not_installed(self):
        with patch("app.services.audit_lab_parsers._HAS_PDFPLUMBER", False):
            with pytest.raises(ImportError, match="pdfplumber"):
                parse_pdf(b"fake")

    @patch("app.services.audit_lab_parsers._HAS_PDFPLUMBER", True)
    @patch("app.services.audit_lab_parsers.pdfplumber")
    def test_pdf_bad_currency_low_confidence(self, mock_pdfplumber: MagicMock):
        table = [
            ["trade_date", "currency_sold", "currency_bought", "amount_sold", "amount_bought"],
            ["2025-06-15", "US", "EUR", "1000000", "920000"],  # "US" not 3-letter
        ]
        mock_pdfplumber.open.return_value = self._mock_pdf([table])

        rows, _, _ = parse_pdf(b"fake")
        conf = row_confidence(rows[0])
        assert conf <= 0.7


# ═══════════════════════════════════════════════════════════════════════════════
# SWIFT MT300/MT320 Parser
# ═══════════════════════════════════════════════════════════════════════════════


class TestSwiftHelpers:
    def test_parse_swift_amount_usd(self):
        ccy, amt = _parse_swift_amount("USD1000000,50")
        assert ccy == "USD"
        assert amt == pytest.approx(1000000.50)

    def test_parse_swift_amount_eur(self):
        ccy, amt = _parse_swift_amount("EUR500000")
        assert ccy == "EUR"
        assert amt == 500000.0

    def test_parse_swift_amount_empty(self):
        ccy, amt = _parse_swift_amount("")
        assert ccy is None
        assert amt is None

    def test_parse_swift_amount_no_match(self):
        ccy, amt = _parse_swift_amount("INVALID")
        assert ccy is None

    def test_extract_tags(self):
        msg = ":20:REF123\n:30T:20250615\n:32B:USD1000000,\n:33B:EUR920000,\n:36:0,92"
        tags = _extract_swift_tags(msg)
        assert tags["20"] == "REF123"
        assert tags["30T"] == "20250615"
        assert tags["32B"] == "USD1000000,"
        assert tags["33B"] == "EUR920000,"
        assert tags["36"] == "0,92"

    def test_split_messages_single(self):
        msg = ":20:REF123\n:30T:20250615"
        parts = _split_swift_messages(msg)
        assert len(parts) == 1

    def test_split_messages_multi(self):
        msg = "{1:header1}\n{4:\n:20:REF1\n:30T:20250101\n-}\n{1:header2}\n{4:\n:20:REF2\n:30T:20250201\n-}"
        parts = _split_swift_messages(msg)
        assert len(parts) >= 2


class TestParseSwiftMT:
    SINGLE_MT300 = (
        ":20:FX-2025-001\n"
        ":30T:20250615\n"
        ":30V:20250617\n"
        ":32B:USD1000000,\n"
        ":33B:EUR920000,\n"
        ":36:0,92\n"
        ":82A:CITIUS33\n"
    )

    def test_basic_mt300(self):
        rows, warnings, pairs = parse_swift_mt(self.SINGLE_MT300)
        assert len(rows) == 1
        row = rows[0]
        assert row["trade_date"] == "2025-06-15"
        assert row["value_date"] == "2025-06-17"
        assert row["currency_sold"] == "USD"
        assert row["currency_bought"] == "EUR"
        assert row["amount_sold"] == 1000000.0
        assert row["amount_bought"] == 920000.0
        assert row["effective_rate"] == pytest.approx(0.92)
        assert row["counterparty"] == "CITIUS33"
        assert row["reference"] == "FX-2025-001"
        assert "USDEUR" in pairs

    def test_confidence_095(self):
        rows, _, _ = parse_swift_mt(self.SINGLE_MT300)
        conf = row_confidence(rows[0])
        assert conf == pytest.approx(0.95)

    def test_mt300_with_fee(self):
        msg = self.SINGLE_MT300 + ":34B:USD500,\n"
        rows, _, _ = parse_swift_mt(msg)
        assert rows[0]["fee_amount"] == 500.0
        assert rows[0]["fee_currency"] == "USD"

    def test_multiple_messages(self):
        msg = (
            "{4:\n"
            ":20:REF1\n"
            ":30T:20250101\n"
            ":32B:USD100000,\n"
            ":33B:EUR92000,\n"
            "-}\n"
            "{4:\n"
            ":20:REF2\n"
            ":30T:20250201\n"
            ":32B:GBP200000,\n"
            ":33B:JPY30000000,\n"
            "-}\n"
        )
        rows, _, pairs = parse_swift_mt(msg)
        assert len(rows) == 2
        assert "USDEUR" in pairs
        assert "GBPJPY" in pairs

    def test_no_tags_generates_warning(self):
        rows, warnings, _ = parse_swift_mt("This is not a SWIFT message at all.")
        assert any("no SWIFT tags" in w for w in warnings)

    def test_partial_tags(self):
        msg = ":20:REF-PARTIAL\n:30T:20250615\n"
        rows, _, _ = parse_swift_mt(msg)
        assert len(rows) == 1
        assert rows[0]["reference"] == "REF-PARTIAL"
        assert rows[0]["trade_date"] == "2025-06-15"
        assert rows[0]["currency_sold"] is None
        assert rows[0]["currency_bought"] is None

    def test_mt320_style(self):
        msg = (
            ":20:DEPO-001\n"
            ":30T:20250615\n"
            ":30V:20250915\n"
            ":32B:USD5000000,\n"
            ":33B:USD5025000,\n"
            ":36:1,005\n"
            ":87A:DEUTDEFF\n"
        )
        rows, _, pairs = parse_swift_mt(msg)
        assert len(rows) == 1
        assert rows[0]["counterparty"] == "DEUTDEFF"
        assert rows[0]["effective_rate"] == pytest.approx(1.005)

    def test_explicit_rate_overrides_computed(self):
        msg = (
            ":20:RATE-TEST\n"
            ":30T:20250615\n"
            ":32B:USD1000000,\n"
            ":33B:EUR920000,\n"
            ":36:0,9150\n"
        )
        rows, _, _ = parse_swift_mt(msg)
        # Tag 36 gives 0.915 which should override the computed 0.92
        assert rows[0]["effective_rate"] == pytest.approx(0.915)


# ═══════════════════════════════════════════════════════════════════════════════
# Cross-parser consistency
# ═══════════════════════════════════════════════════════════════════════════════


class TestCrossParserConsistency:
    """All parsers must return the same tuple shape and row keys."""

    EXPECTED_KEYS = {
        "row_index", "trade_date", "value_date", "currency_sold",
        "currency_bought", "amount_sold", "amount_bought", "effective_rate",
        "counterparty", "fee_amount", "fee_currency", "reference",
        "parse_warnings",
    }

    def test_csv_row_keys(self):
        raw = b"trade_date,currency_sold,currency_bought,amount_sold,amount_bought\n2025-01-01,USD,EUR,100,92\n"
        rows, _, _ = parse_csv(raw)
        assert set(rows[0].keys()) == self.EXPECTED_KEYS

    def test_xlsx_row_keys(self):
        import openpyxl as xl

        wb = xl.Workbook()
        ws = wb.active
        ws.append(["trade_date", "currency_sold", "currency_bought", "amount_sold", "amount_bought"])
        ws.append(["2025-01-01", "USD", "EUR", "100", "92"])
        buf = io.BytesIO()
        wb.save(buf)

        rows, _, _ = parse_xlsx(buf.getvalue())
        assert set(rows[0].keys()) == self.EXPECTED_KEYS

    @patch("app.services.audit_lab_parsers._HAS_PDFPLUMBER", True)
    @patch("app.services.audit_lab_parsers.pdfplumber")
    def test_pdf_row_keys(self, mock_pdfplumber: MagicMock):
        table = [
            ["trade_date", "currency_sold", "currency_bought", "amount_sold", "amount_bought"],
            ["2025-01-01", "USD", "EUR", "100", "92"],
        ]
        mock_pdf = MagicMock()
        page = MagicMock()
        page.extract_tables.return_value = [table]
        mock_pdf.pages = [page]
        mock_pdf.close = MagicMock()
        mock_pdfplumber.open.return_value = mock_pdf

        rows, _, _ = parse_pdf(b"fake")
        assert set(rows[0].keys()) == self.EXPECTED_KEYS

    def test_swift_row_keys(self):
        msg = ":20:REF\n:30T:20250101\n:32B:USD100,\n:33B:EUR92,\n"
        rows, _, _ = parse_swift_mt(msg)
        assert set(rows[0].keys()) == self.EXPECTED_KEYS

    def test_return_types(self):
        """All parsers return tuple[list[dict], list[str], set[str]]."""
        raw = b"trade_date,currency_sold,currency_bought,amount_sold,amount_bought\n2025-01-01,USD,EUR,100,92\n"
        rows, warnings, pairs = parse_csv(raw)
        assert isinstance(rows, list)
        assert isinstance(warnings, list)
        assert isinstance(pairs, set)
        for w in warnings:
            assert isinstance(w, str)
        for p in pairs:
            assert isinstance(p, str)
